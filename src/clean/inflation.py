"""
Inflation CPI cleaning utilities
Processes inflation CPI data for the same 32 countries used in the main analysis.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

# Import from centralized constants
from .constants import TARGET_ISO3_32
from .utils import (
    filter_to_target_countries,
    filter_to_year_range,
    save_dataframe,
)
from .worldbank import WorldBankProcessor


def read_inflation_excel(path: str | Path, sheet_name: str | int = 0) -> pd.DataFrame:
    """
    Read Inflation CPI Excel file with extreme robustness.
    Includes smart header detection to skip metadata rows.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path.resolve()}")

    # helper for smart header detection
    def find_header_row(data_rows):
        keywords = {"country", "reference area", "ref_area", "iso3", "location"}
        for i, row in enumerate(data_rows[:20]):
            row_str = [str(c).lower().strip() for c in row if c is not None]
            if any(k in row_str for k in keywords):
                return i
        return 0

    # 1. Try Calamine first (most robust against style corruption)
    try:
        try:
            from python_calamine import CalamineWorkbook
        except ImportError:
            from calamine import CalamineWorkbook

        workbook = CalamineWorkbook.from_path(str(path))
        sheet_names = workbook.sheet_names
        target = sheet_names[sheet_name] if isinstance(sheet_name, int) else sheet_name
        data = workbook.get_sheet_by_name(target).to_python()

        if data:
            header_idx = find_header_row(data)
            df = pd.DataFrame(data[header_idx + 1 :], columns=data[header_idx])
            return df
    except Exception as e:
        print(f"⚠️  Calamine read failed: {e}")

    # 2. Try standard pandas read with manual header detection
    try:
        # Load a few rows to find header
        preview = pd.read_excel(path, sheet_name=sheet_name, nrows=20, header=None)
        data_preview = preview.values.tolist()
        header_idx = find_header_row(data_preview)

        # Reload with correct header
        return pd.read_excel(path, sheet_name=sheet_name, header=header_idx)
    except Exception as e:
        error_msg = str(e)
        print(f"⚠️  Pandas read failed: {error_msg}")

    # 3. CRITICAL FALLBACK: Style Stripping (if it's a style error)
    if "expected" in error_msg and "Fill" in error_msg:
        import tempfile
        import zipfile

        print("🛠️  Corrupted Excel styles detected. Cleaning file metadata...")
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_path = Path(tmpdir) / "cleaned_data.xlsx"
                with zipfile.ZipFile(path, "r") as zin:
                    with zipfile.ZipFile(tmp_path, "w") as zout:
                        for item in zin.infolist():
                            if item.filename != "xl/styles.xml":
                                zout.writestr(item, zin.read(item.filename))

                # Try reading the cleaned version with header detection
                preview = pd.read_excel(
                    tmp_path, sheet_name=sheet_name, nrows=20, header=None, engine="openpyxl"
                )
                header_idx = find_header_row(preview.values.tolist())
                return pd.read_excel(
                    tmp_path, sheet_name=sheet_name, header=header_idx, engine="openpyxl"
                )
        except Exception as clean_e:
            print(f"❌ Style cleaning failed: {clean_e}")

    # 4. Final attempt: Manual openpyxl
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        data = [row for row in ws.iter_rows(values_only=True)]
        if data:
            header_idx = find_header_row(data)
            return pd.DataFrame(data[header_idx + 1 :], columns=data[header_idx])
    except Exception as final_e:
        print(f"❌ All extraction methods failed for {path.name}")
        raise final_e

    # 4. Final attempt: Manual openpyxl with values_only=True
    try:
        print("🔍 Attempting manual extraction (ignoring metadata)...")
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        data = [row for row in ws.iter_rows(values_only=True)]
        if data:
            return pd.DataFrame(data[1:], columns=data[0])
    except Exception as final_e:
        print(f"❌ All extraction methods failed for {path.name}")
        raise final_e


def standardize_inflation_to_long(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Convert inflation data from wide to long format.
    Handles World Bank-style format with years as columns.
    Returns: country, year, inflation_cpi
    """
    return WorldBankProcessor.wide_to_long(df_raw, value_name="inflation_cpi")


# Removed: Now using shared map_country_to_iso3 from utils.py


def filter_32_countries(
    df_mapped: pd.DataFrame,
    year_min: int | None = 1980,
    year_max: int | None = 2023,
    target_iso3: set[str] = TARGET_ISO3_32,
) -> pd.DataFrame:
    """
    Filter inflation data to:
    - The same 32 countries used in the analysis
    - Year range (default 1980-2023)
    - Keep only: iso3, year, inflation_cpi
    """
    # Use shared utility functions
    out = filter_to_year_range(df_mapped, year_min, year_max)
    out = filter_to_target_countries(out, target_iso3)

    # Keep only essential columns
    out = out[["iso3", "year", "inflation_cpi"]].copy()

    # Sort by iso3 and year
    out = out.sort_values(["iso3", "year"]).reset_index(drop=True)

    return out


def save_inflation(df: pd.DataFrame, out_path: str | Path) -> Path:
    """Save processed inflation data to parquet or CSV."""
    return save_dataframe(df, out_path)
